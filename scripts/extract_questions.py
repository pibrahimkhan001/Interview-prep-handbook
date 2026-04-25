"""Extract all questions from the Interview Prep Excel into a structured JSON."""
import json
import re
from openpyxl import load_workbook

SRC = "/mnt/project/Ibrahim_Interview_Prep.xlsx"
OUT = "/home/claude/interview-prep-guide/data/questions.json"


def clean_question(text):
    """Strip leading numbers/whitespace/tabs from a question line."""
    if not text:
        return ""
    # Remove leading "N\t" or "N    " or "N)" etc.
    text = re.sub(r"^\s*\d+[\t\s\.\)]+", "", text.strip())
    return text.strip()


def parse_cell(cell_text):
    """Split a cell's multi-line text into individual questions."""
    if not cell_text:
        return []
    lines = [ln for ln in cell_text.split("\n") if ln.strip()]
    questions = []
    for ln in lines:
        q = clean_question(ln)
        if q and len(q) > 5:  # filter out junk
            questions.append(q)
    return questions


def main():
    wb = load_workbook(SRC)
    ws = wb["Interview Practice Questions"]

    # Row 1 = topic headers (col 1..13)
    topics = []
    for col_idx in range(2, 15):  # B..N
        val = ws.cell(row=1, column=col_idx).value
        if val:
            topics.append(val.strip())

    # Rows 2..8 = difficulty levels
    levels = []
    data = {}
    for row_idx in range(2, 10):
        level = ws.cell(row=row_idx, column=1).value
        if not level:
            continue
        level = level.strip()
        levels.append(level)
        data[level] = {}
        for topic_idx, topic in enumerate(topics):
            cell_val = ws.cell(row=row_idx, column=topic_idx + 2).value
            questions = parse_cell(cell_val) if cell_val else []
            data[level][topic] = questions

    # Reorganize by topic -> level -> questions
    by_topic = {}
    for topic in topics:
        by_topic[topic] = {}
        for level in levels:
            qs = data.get(level, {}).get(topic, [])
            if qs:
                by_topic[topic][level] = qs

    # Stats
    total = 0
    print(f"Topics: {len(topics)}")
    print(f"Levels: {levels}\n")
    for topic in topics:
        topic_total = sum(len(qs) for qs in by_topic.get(topic, {}).values())
        total += topic_total
        levels_covered = list(by_topic.get(topic, {}).keys())
        print(f"  {topic}: {topic_total} questions across {len(levels_covered)} levels")
    print(f"\nGRAND TOTAL: {total} questions")

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump({"topics": topics, "levels": levels, "data": by_topic}, f, indent=2, ensure_ascii=False)
    print(f"\nWrote: {OUT}")


if __name__ == "__main__":
    main()
